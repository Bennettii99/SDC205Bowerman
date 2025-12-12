# Final Performance Assessment - Ishmael Bennett
print("Ishen9200")

import csv
import os
from datetime import date
import openpyxl
from openpyxl.chart import PieChart, Reference
import matplotlib.pyplot as plt

# Change to the FinalExam folder
os.chdir(r"C:\FinalExam")

# Task 1: Ask for 5 numbers and show the sum
def askUser():
    print("\nPlease enter a number: ", end="")
    total = 0
    for i in range(5):
        num = int(input("Please enter a number: "))
        total += num
    print(f"The sum for the 5 numbers entered is: {total}")

# Task 2: Ask for 5 names & incomes and append to final.csv
def askIncome():
    # This loop asks for 5 people and adds them to final.csv - exact match to instructor
    with open("final.csv", "a", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        for i in range(5):
            name = input("Please enter a name: ")
            income = input("Please enter their income: ")
            writer.writerow([name, income])

# Task 3: Create pie chart in Excel (final.xlsx)
def excelPie():
    # Load the updated CSV data
    names = []
    incomes = []
    with open("final.csv", "r") as file:
        reader = csv.reader(file)
        for row in reader:
            names.append(row[0])
            incomes.append(int(row[1]))   # convert to int or pie chart is blank!!

    # Create new workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income Data"

    # Write headers and data
    ws['A1'] = "Name"
    ws['B1'] = "Income"
    for i in range(len(names)):
        ws.cell(row=i+2, column=1).value = names[i]
        ws.cell(row=i+2, column=2).value = incomes[i]

    # Create pie chart
    data = Reference(ws, min_col=2, min_row=1, max_row=len(names)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(names)+1)
    pie = PieChart()
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = f"Ishben9200 {date.today():%B %d, %Y}"   

    ws.add_chart(pie, "A10")
    wb.save("final.xlsx")

# Task 4: Create vertical bar chart using matplotlib
def verticalBar():
    names = []
    incomes = []
    with open("final.csv", "r") as file:
        reader = csv.reader(file)
        for row in reader:
            names.append(row[0])
            incomes.append(int(row[1]))

    plt.figure(figsize=(10, 6))
    plt.bar(names, incomes, color='green')
    plt.title(f"Ishben9200 {date.today():%B %d, %Y}")  
    plt.xlabel("Name")
    plt.ylabel("Income")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

# Run everything
askUser()
askIncome()
excelPie()
verticalBar()