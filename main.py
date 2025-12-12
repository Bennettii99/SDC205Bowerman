"""
Student Name: [Your Name]
Student ID: [Your Student ID]
Date: December 11, 2025

This program reads measurement data from a CSV file (weight in pounds, height in inches, or temperature in Fahrenheit),
allows conversion and charting using openpyxl. Option 3 generates a report with a bar or line chart saved to final.xlsx.
"""

import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# Example CSV structure expected:
# Date,Value
# 2025-01-01,180
# 2025-01-02,182
# etc.

def createChart(csv_file_path: str, chart_type: str):
    """
    Creates an Excel chart (bar or line) from CSV data and saves it to final.xlsx

    Args:
        csv_file_path (str): Path to the input CSV file containing Date and Value columns
        chart_type (str): Type of chart to create - either "bar" or "line"

    Returns:
        None: Saves final.xlsx with chart and data
    """
    # Ask user whether to use original or converted data
    print("\nChoose data source for the chart:")
    print("1. Original data (pounds / inches / Fahrenheit)")
    print("2. Converted data (kilograms / centimeters / Celsius)")
    choice = input("Enter 1 or 2: ").strip()

    is_converted = (choice == "2")

    # Determine unit labels based on choice
    if "weight" in csv_file_path.lower() or "pounds" in csv_file_path.lower():
        original_unit = "Pounds"
        converted_unit = "Kilograms"
        y_label = converted_unit if is_converted else original_unit
        conversion_factor = 0.453592  # pounds to kg
    elif "height" in csv_file_path.lower() or "inches" in csv_file_path.lower():
        original_unit = "Inches"
        converted_unit = "Centimeters"
        y_label = converted_unit if is_converted else original_unit
        conversion_factor = 2.54  # inches to cm
    else:  # Assume temperature
        original_unit = "Fahrenheit"
        converted_unit = "Celsius"
        y_label = converted_unit if is_converted else original_unit
        def f_to_c(f): return (f - 32) * 5/9
        conversion_factor = None  # Special case

    # Read data from CSV
    dates = []
    values = []

    with open(csv_file_path, newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            dates.append(row['Date'])
            value = float(row['Value'])
            if is_converted:
                if conversion_factor is not None:
                    value *= conversion_factor
                else:
                    value = f_to_c(value)
            values.append(round(value, 2))

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Measurement Data"

    # Write headers and data
    ws['A1'] = "Date"
    ws['B1'] = y_label

    for i, (date, value) in enumerate(zip(dates, values), start=2):
        ws[f'A{i}'] = date
        ws[f'B{i}'] = value

    # Create chart
    if chart_type == "bar":
        chart = BarChart()
    else:
        chart = LineChart()

    # Add data to chart
    data = Reference(ws, min_col=2, min_row=1, max_row=len(values)+1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(values)+1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Customize chart
    chart.title = f"{input('\nEnter your Student ID: ')} {datetime.now().strftime('%m/%d/%Y')}"
    chart.x_axis.title = "Date"
    chart.y_axis.title = y_label
    chart.height = 12
    chart.width = 20

    # Add chart to worksheet
    ws.add_chart(chart, "D2")

    # Save the workbook
    wb.save("final.xlsx")
    print(f"\nChart created successfully! Saved as 'final.xlsx' with {chart_type} chart using {y_label}.")


def generateReport(csv_file_path: str):
    """
    Prompts user for chart type and calls createChart to generate the report

    Args:
        csv_file_path (str): Path to the CSV file containing the measurement data

    Returns:
        None: Calls createChart() which generates final.xlsx
    """
    print("\n--- Generate Report ---")
    print("Select chart type:")
    print("1. Bar Chart")
    print("2. Line Chart")
    choice = input("Enter 1 or 2: ").strip()

    if choice == "1":
        createChart(csv_file_path, "bar")
    elif choice == "2":
        createChart(csv_file_path, "line")
    else:
        print("Invalid choice. Please try again.")
        generateReport(csv_file_path)


# === Main Program (Updated Menu) ===
def main():
    # Update this path to your actual CSV file
    CSV_FILE_PATH = "measurements1.csv"  # Make sure this file exists!

    while True:
        print("\n" + "="*50)
        print("MAIN MENU")
        print("="*50)
        print("1. Enter new data")
        print("2. View previous data")
        print("3. Generate Report")   # NEW OPTION
        print("4. Exit")
        print("="*50)

        choice = input("Enter your choice (1-4): ").strip()

        if choice == "1":
            print("Feature not implemented in this version.")
            # (You would normally have data entry here)
        elif choice == "2":
            print("Feature not implemented in this version.")
            # (You would normally display data here)
        elif choice == "3":
            try:
                generateReport(CSV_FILE_PATH)
            except FileNotFoundError:
                print(f"Error: Could not find {CSV_FILE_PATH}. Please ensure the file exists.")
            except Exception as e:
                print(f"An error occurred: {e}")
        elif choice == "4":
            print("Goodbye!")
            break
        else:
            print("Invalid choice. Please enter 1, 2, 3, or 4.")


if __name__ == "__main__":
    main()